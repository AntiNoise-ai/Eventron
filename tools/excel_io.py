"""Excel import/export for attendee lists and seat maps.

Pure functions — take data in, return data out. No DB access.
"""

from io import BytesIO
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook


# ── Column mapping: Excel header → attendee field ────────────
DEFAULT_COLUMN_MAP = {
    "姓名": "name",
    "name": "name",
    "职位": "title",
    "title": "title",
    "公司": "organization",
    "organization": "organization",
    "部门": "department",
    "department": "department",
    "角色": "role",
    "role": "role",
    "电话": "phone",
    "phone": "phone",
    "邮箱": "email",
    "email": "email",
}


def import_attendees_from_excel(
    file_path: Path | None = None,
    file_bytes: bytes | None = None,
    column_map: dict[str, str] | None = None,
) -> list[dict[str, Any]]:
    """Parse an Excel file into a list of attendee dicts.

    Args:
        file_path: Path to .xlsx file (mutually exclusive with file_bytes).
        file_bytes: Raw bytes of .xlsx file.
        column_map: Custom header→field mapping. Falls back to DEFAULT_COLUMN_MAP.

    Returns:
        List of attendee dicts with normalized field names.

    Raises:
        ValueError: If neither file_path nor file_bytes is provided,
                    or if no valid rows found.
    """
    if file_path is None and file_bytes is None:
        raise ValueError("Provide either file_path or file_bytes")

    cmap = {k.lower().strip(): v for k, v in (column_map or DEFAULT_COLUMN_MAP).items()}

    if file_bytes:
        wb = load_workbook(BytesIO(file_bytes), read_only=True)
    else:
        wb = load_workbook(file_path, read_only=True)

    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if len(rows) < 2:
        raise ValueError("Excel file must have a header row and at least one data row")

    # Map headers
    headers = [str(h).lower().strip() if h else "" for h in rows[0]]
    field_indices: dict[str, int] = {}
    for i, header in enumerate(headers):
        if header in cmap:
            field_indices[cmap[header]] = i

    if "name" not in field_indices:
        raise ValueError("Excel file must have a '姓名' or 'name' column")

    # Parse data rows
    attendees = []
    for row in rows[1:]:
        if not row or not any(row):
            continue
        att: dict[str, Any] = {"attrs": {}}
        for field, idx in field_indices.items():
            val = row[idx] if idx < len(row) else None
            att[field] = str(val).strip() if val is not None else None
        if att.get("name"):
            attendees.append(att)

    return attendees


def export_attendees_to_excel(
    attendees: list[dict[str, Any]],
    seats: list[dict[str, Any]] | None = None,
) -> bytes:
    """Export attendee list (optionally with seat info) to Excel bytes.

    Args:
        attendees: List of attendee dicts.
        seats: Optional list of seat dicts (matched by attendee_id).

    Returns:
        Bytes of the generated .xlsx file.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "参会人员"

    # Build seat lookup
    seat_map: dict[str, dict] = {}
    if seats:
        for s in seats:
            if s.get("attendee_id"):
                seat_map[s["attendee_id"]] = s

    # Headers
    headers = ["姓名", "职位", "公司", "部门", "角色", "电话", "邮箱", "状态"]
    if seats:
        headers.extend(["座位号", "排", "列"])
    ws.append(headers)

    # Data rows
    for att in attendees:
        row = [
            att.get("name", ""),
            att.get("title", ""),
            att.get("organization", ""),
            att.get("department", ""),
            att.get("role", ""),
            att.get("phone", ""),
            att.get("email", ""),
            att.get("status", ""),
        ]
        if seats:
            seat = seat_map.get(att.get("id", ""), {})
            row.extend([
                seat.get("label", ""),
                seat.get("row_num", ""),
                seat.get("col_num", ""),
            ])
        ws.append(row)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_seatmap_to_excel(
    seats: list[dict[str, Any]],
    rows: int,
    cols: int,
) -> bytes:
    """Export a visual seat map grid to Excel.

    Each cell shows the attendee name or '空' for empty seats.

    Args:
        seats: List of seat dicts with row_num, col_num, and optional attendee_name.
        rows: Total rows in venue.
        cols: Total cols in venue.

    Returns:
        Bytes of the generated .xlsx file.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "座位图"

    # Build lookup
    seat_map: dict[tuple[int, int], dict] = {}
    for s in seats:
        seat_map[(s["row_num"], s["col_num"])] = s

    # Header row (column numbers)
    ws.append([""] + [f"列{c}" for c in range(1, cols + 1)])

    # Data rows
    for r in range(1, rows + 1):
        row = [f"第{r}排"]
        for c in range(1, cols + 1):
            s = seat_map.get((r, c), {})
            if s.get("seat_type") in ("disabled", "aisle"):
                row.append("—")
            elif s.get("attendee_name"):
                row.append(s["attendee_name"])
            else:
                row.append("空")
        ws.append(row)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
